"""
Leitor tolerante do "Suporte IES.xlsx".

Suporta dois formatos:
  A) uma aba por empresa   (formato atual) -> grupo vem do nome da aba
  B) aba unica com coluna "Company"        (formato antigo) -> grupo vem da coluna

Cada aba pode ter um schema proprio. As colunas de codigo e de nome sao detectadas
por heuristica sobre o cabecalho, entao nao e preciso padronizar as planilhas de origem.
"""
import os
import re
import unicodedata

import openpyxl
import pandas as pd


def norm(s):
    """Maiuscula, sem acento, sem pontuacao redundante, espacos normalizados."""
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace(".", " ").replace("_", " ")
    return re.sub(r"\s+", " ", s).upper().strip()


def _e_coluna_codigo(h):
    """Cabecalho de codigo de IES? Exclui explicitamente codigo de mantenedora."""
    if "MANTENEDORA" in h or "CNPJ" in h:
        return False
    if h in ("IES CODE", "CODIGO IES", "COD IES"):
        return True
    tem_cod = "COD" in h or "CODIGO" in h
    return tem_cod and ("IES" in h or "MEC" in h)


def _e_coluna_nome(h):
    # nunca confundir com a coluna de codigo (ex.: cabecalho "IES Code")
    if _e_coluna_codigo(h) or "CODIGO" in h or "COD " in h:
        return False
    if "MANTENEDORA" in h or "MUNICIPIO" in h or "CNPJ" in h:
        return False
    if h in ("UF", "ESTADO", "SIGLA", "CIDADE", "CITY", "STATE", "REGIAO", "SITUACAO", "MARCA"):
        return False
    # ex.: "IES", "MANTIDA", "UNIDADE", "IES/Mantidas",
    #      "Instituicao de Ensino Superior", "NOME DA MANTIDA (IES)"
    return h in ("IES", "MANTIDA", "MANTIDAS", "UNIDADE") or \
        "MANTIDA" in h or "INSTITUICAO" in h or "NOME" in h


def _acha_cabecalho(ws, max_scan=6):
    """Retorna (indice_linha_cabecalho, headers_normalizados) ou (None, None)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        hs = [norm(c) for c in row]
        if any(_e_coluna_codigo(h) for h in hs):
            return i, hs
    return None, None


def ler_suporte(caminho, mapa_abas=None):
    """
    Le o Suporte IES.xlsx e devolve (DataFrame, avisos).

    DataFrame: CO_IES | GRUPO | NO_IES_SUPORTE | ABA
    """
    mapa_abas = mapa_abas or {}
    wb = openpyxl.load_workbook(caminho, data_only=True)
    registros, avisos = [], []

    for ws in wb.worksheets:
        aba = ws.title.strip()
        lin, hs = _acha_cabecalho(ws)
        if lin is None:
            avisos.append(f"aba '{aba}': nenhuma coluna de codigo de IES reconhecida — ignorada")
            continue

        i_cod = next(i for i, h in enumerate(hs) if _e_coluna_codigo(h))
        i_nome = next((i for i, h in enumerate(hs) if _e_coluna_nome(h)), None)
        i_grupo = next((i for i, h in enumerate(hs) if h in ("COMPANY", "GRUPO", "EMPRESA")), None)
        # colunas opcionais, quando a aba as tiver
        i_marca = next((i for i, h in enumerate(hs) if h in ("MARCA", "BANDEIRA", "BRAND")), None)
        i_sit = next((i for i, h in enumerate(hs) if h in ("SITUACAO", "STATUS")), None)

        # grupo da aba: mapa explicito > nome da aba
        chave = norm(aba)
        grupo_aba = mapa_abas.get(chave)
        if grupo_aba is None and i_grupo is None:
            grupo_aba = aba
            avisos.append(f"aba '{aba}': nao esta em config/suporte_abas.csv — usando '{aba}' como grupo")

        n_aba = 0
        for row in ws.iter_rows(min_row=lin + 1, values_only=True):
            if i_cod >= len(row):
                continue
            cod = row[i_cod]
            if cod is None or str(cod).strip() == "":
                continue
            try:
                cod = int(float(str(cod).strip()))
            except ValueError:
                continue

            grupo = grupo_aba
            if i_grupo is not None and i_grupo < len(row) and row[i_grupo]:
                grupo = str(row[i_grupo]).strip()
            if not grupo:
                continue

            nome = ""
            if i_nome is not None and i_nome < len(row) and row[i_nome]:
                nome = re.sub(r"\s+", " ", str(row[i_nome])).strip()

            def _opt(idx):
                if idx is None or idx >= len(row) or row[idx] is None:
                    return ""
                return re.sub(r"\s+", " ", str(row[idx])).strip()

            registros.append({"CO_IES": cod, "GRUPO": grupo, "NO_IES_SUPORTE": nome, "ABA": aba,
                              "MARCA": _opt(i_marca), "SITUACAO": _opt(i_sit)})
            n_aba += 1

        if n_aba == 0:
            avisos.append(f"aba '{aba}': cabecalho encontrado mas nenhuma linha de IES lida")

    df = pd.DataFrame(registros)
    if df.empty:
        return df, avisos

    # conflito: mesmo CO_IES em dois grupos
    conf = df.groupby("CO_IES")["GRUPO"].nunique()
    for co in conf[conf > 1].index:
        gs = sorted(df.loc[df["CO_IES"] == co, "GRUPO"].unique())
        avisos.append(f"CO_IES {co} aparece em mais de um grupo: {gs} — mantido o primeiro")

    df["CO_IES"] = df["CO_IES"].astype("int64")
    return df, avisos


# --------------------------------------------------------------------------
# Desambiguacao codigo de IES x codigo de mantenedora
# --------------------------------------------------------------------------
_STOP = {"DE", "DA", "DO", "DAS", "DOS", "E", "LTDA", "SA", "S", "A", "ME", "EPP",
         "EIRELI", "SS", "SOCIEDADE", "CENTRO", "FACULDADE", "INSTITUTO", "ASSOCIACAO"}


def _tokens(s):
    return {t for t in re.split(r"[^A-Z0-9]+", norm(s)) if t and t not in _STOP and len(t) > 1}


def _similar(a, b):
    """0..1. Combina razao de sequencia com sobreposicao de tokens significativos."""
    from difflib import SequenceMatcher
    na, nb = norm(a), norm(b)
    if not na or not nb:
        return 0.0
    seq = SequenceMatcher(None, na, nb).ratio()
    ta, tb = _tokens(a), _tokens(b)
    if not ta or not tb:
        return seq
    cont = len(ta & tb) / min(len(ta), len(tb))   # containment, tolera nome truncado
    return max(seq, cont)


def classifica_codigos(df, censo_ies, limiar=0.55, margem=0.12, piso_mant=0.45):
    """
    Decide, por linha, se o codigo informado e de IES ou de MANTENEDORA, comparando o
    nome da planilha com o nome da IES e o nome da mantenedora no Censo.

    Necessario porque algumas abas misturam blocos de IES e de mantenedoras na mesma
    coluna, e ha codigos que existem nos dois universos apontando para entidades diferentes.

    censo_ies: DataFrame com CO_IES, NO_IES, CO_MANTENEDORA, NO_MANTENEDORA.
    Devolve df com TIPO_CODIGO ('ies' | 'mantenedora' | 'incerto'), SCORE_IES, SCORE_MANT.
    """
    nome_ies = dict(zip(censo_ies["CO_IES"], censo_ies["NO_IES"]))
    nome_mant = (censo_ies.dropna(subset=["CO_MANTENEDORA"])
                 .drop_duplicates("CO_MANTENEDORA")
                 .set_index("CO_MANTENEDORA")["NO_MANTENEDORA"].to_dict())

    tipos, s_ies, s_mant, revisar = [], [], [], []
    for r in df.itertuples():
        co, nome = r.CO_IES, r.NO_IES_SUPORTE
        existe_ies, existe_mant = co in nome_ies, co in nome_mant
        si = _similar(nome, nome_ies[co]) if (existe_ies and nome) else 0.0
        sm = _similar(nome, nome_mant[co]) if (existe_mant and nome) else 0.0
        rev = False

        if not nome:
            # sem nome nao da para desambiguar: assume IES (interpretacao padrao da coluna)
            tipo = "ies" if existe_ies or not existe_mant else "mantenedora"
            rev = existe_ies and existe_mant
        elif sm >= limiar and sm > si + margem:
            tipo = "mantenedora"
        elif si >= limiar:
            tipo = "ies"
        elif existe_ies and not existe_mant:
            tipo = "ies"
        elif existe_mant and not existe_ies:
            tipo = "mantenedora"
        elif existe_ies and existe_mant:
            # Ambiguo. A coluna foi rotulada como codigo de IES, entao "ies" e o padrao:
            # so vira mantenedora se houver evidencia minima de nome a favor disso.
            # Evita decidir por empates fracos (ex.: 0,23 vs 0,24).
            tipo = "mantenedora" if (sm >= piso_mant and sm > si + margem) else "ies"
            rev = True
        else:
            tipo = "ies"          # codigo desconhecido: IES ausente do Censo
        tipos.append(tipo)
        s_ies.append(round(si, 3))
        s_mant.append(round(sm, 3))
        revisar.append(rev)

    out = df.copy()
    out["TIPO_CODIGO"] = tipos
    out["SCORE_IES"] = s_ies
    out["SCORE_MANT"] = s_mant
    out["REVISAR"] = revisar
    return out


def ler_csv_comentado(path, sep=";"):
    """Le CSV ignorando linhas iniciadas por '#'. Devolve DataFrame vazio se nao houver dados."""
    if not os.path.exists(path):
        return pd.DataFrame()
    linhas = [ln for ln in open(path, encoding="utf-8-sig").read().splitlines()
              if ln.strip() and not ln.lstrip().startswith("#")]
    if len(linhas) <= 1:
        return pd.DataFrame()
    from io import StringIO
    return pd.read_csv(StringIO("\n".join(linhas)), sep=sep)
